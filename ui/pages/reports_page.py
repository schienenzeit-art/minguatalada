"""
Berichte-Seite – vollständig native PyQt6-Implementierung, keine WebEngine.

Tabs:
  1. Übersicht      – Anträge/Karten nach Status + KPI-Kacheln
  2. Jahresauswertung – 12-Monats-Tabelle mit Vorjahresvergleich
  3. Zeitraum        – Von-bis-Auswertung mit Monatsaufschlüsselung
  4. Warteliste      – offene Fälle nach Wartezeit
"""
from datetime import date

from PyQt6.QtCore import Qt, QUrl
from PyQt6.QtGui import QColor, QFont, QDesktopServices
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QTabWidget, QGroupBox, QDateEdit, QSpinBox, QFrame,
    QScrollArea, QCheckBox, QSizePolicy, QFileDialog, QMessageBox,
)
from PyQt6.QtCore import QDate

from core.session import Session
from services.report_service import ReportService
from ui.components.page_header import PageHeader

_BOLD_FONT = QFont()
_BOLD_FONT.setBold(True)

_COL_TOTALS_BG  = QColor("#e8f4fd")
_COL_APPROVED   = QColor("#e8f8ed")
_COL_REJECTED   = QColor("#fdeaea")
_COL_HARDSHIP   = QColor("#fff3e0")
_COL_HEADER_FG  = QColor("#1a5c37")


def _make_item(text: str, align_right: bool = False, bold: bool = False,
               bg: QColor | None = None, fg: QColor | None = None) -> QTableWidgetItem:
    item = QTableWidgetItem(str(text))
    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
    if align_right:
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
    if bold:
        item.setFont(_BOLD_FONT)
    if bg:
        item.setBackground(bg)
    if fg:
        item.setForeground(fg)
    return item


def _pct(part: int, total: int) -> str:
    if not total:
        return "–"
    return f"{round(part / total * 100, 1)} %"


def _delta_str(current: int, previous: int) -> str:
    if previous == 0:
        return "–"
    diff = current - previous
    sign = "+" if diff >= 0 else ""
    pct  = round(diff / previous * 100, 1)
    return f"{sign}{diff} ({sign}{pct} %)"


class ReportsPage(QWidget):
    def __init__(self, report_service: ReportService | None = None):
        super().__init__()
        self.report_service = report_service or ReportService()
        self._locations: list[dict] = []
        self.setup_ui()
        self._load_locations()

    # ── Haupt-UI ──────────────────────────────────────────────────────────────

    def setup_ui(self):
        self.setObjectName("reportsPage")
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 24, 24, 24)
        root.setSpacing(16)

        root.addWidget(PageHeader(
            title="Berichte",
            subtitle="Auswertungen nach Standort, Zeitraum und Kennzahl.",
        ))

        # ── Globaler Standort-Filter + Export-Leiste ──────────────────────────
        loc_bar = QHBoxLayout()
        loc_bar.setSpacing(10)
        loc_bar.addWidget(QLabel("Standort:"))
        self._location_combo = QComboBox()
        self._location_combo.setMinimumWidth(220)
        loc_bar.addWidget(self._location_combo)
        loc_bar.addStretch()

        btn_pdf = QPushButton("PDF exportieren")
        btn_pdf.setObjectName("SoftButton")
        btn_pdf.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_pdf.setToolTip("Aktuelle Auswertung als PDF exportieren")
        btn_pdf.clicked.connect(self._export_to_pdf)
        loc_bar.addWidget(btn_pdf)

        btn_excel = QPushButton("Excel exportieren")
        btn_excel.setObjectName("SoftButton")
        btn_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_excel.setToolTip("Aktuelle Auswertung als Excel-Datei exportieren")
        btn_excel.clicked.connect(self._export_to_excel)
        loc_bar.addWidget(btn_excel)

        root.addLayout(loc_bar)

        # ── Tabs ──────────────────────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.addTab(self._build_overview_tab(),    "Übersicht")
        self._tabs.addTab(self._build_annual_tab(),      "Jahresauswertung")
        self._tabs.addTab(self._build_period_tab(),      "Zeitraum")
        self._tabs.addTab(self._build_waitlist_tab(),    "Warteliste")
        root.addWidget(self._tabs, 1)

    # ── Tab 1: Übersicht ──────────────────────────────────────────────────────

    def _build_overview_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        # Filter-Zeile
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Von:"))
        self._ov_from = QDateEdit()
        self._ov_from.setCalendarPopup(True)
        self._ov_from.setDate(QDate(date.today().year, 1, 1))
        self._ov_from.setDisplayFormat("dd.MM.yyyy")
        filter_row.addWidget(self._ov_from)
        filter_row.addWidget(QLabel("Bis:"))
        self._ov_to = QDateEdit()
        self._ov_to.setCalendarPopup(True)
        self._ov_to.setDate(QDate.currentDate())
        self._ov_to.setDisplayFormat("dd.MM.yyyy")
        filter_row.addWidget(self._ov_to)
        self._ov_all_time = QCheckBox("Gesamter Zeitraum (kein Datumsfilter)")
        self._ov_all_time.setChecked(True)
        self._ov_all_time.stateChanged.connect(self._on_ov_alltime_changed)
        filter_row.addWidget(self._ov_all_time)
        filter_row.addStretch()
        btn_ov = QPushButton("Auswertung starten")
        btn_ov.setObjectName("PrimaryButton")
        btn_ov.clicked.connect(self._run_overview)
        filter_row.addWidget(btn_ov)
        layout.addLayout(filter_row)
        self._on_ov_alltime_changed()

        # KPI-Kacheln
        self._kpi_row = QHBoxLayout()
        self._kpi_row.setSpacing(12)
        self._kpi_total_claims  = self._make_kpi("Anträge gesamt", "–")
        self._kpi_total_cards   = self._make_kpi("Karten aktiv", "–")
        self._kpi_total_persons = self._make_kpi("Personen", "–")
        self._kpi_open          = self._make_kpi("In Prüfung", "–")
        for kpi in (self._kpi_total_claims, self._kpi_total_cards,
                    self._kpi_total_persons, self._kpi_open):
            self._kpi_row.addWidget(kpi)
        self._kpi_row.addStretch()
        layout.addLayout(self._kpi_row)

        # Tabellen
        tables_row = QHBoxLayout()
        tables_row.setSpacing(16)

        left = QVBoxLayout()
        left.addWidget(QLabel("<b>Anträge nach Status</b>"))
        self._ov_claim_table = self._make_table(["Status", "Anzahl", "Anteil"])
        left.addWidget(self._ov_claim_table)
        tables_row.addLayout(left, 3)

        right = QVBoxLayout()
        right.addWidget(QLabel("<b>Bezugskarten nach Status</b>"))
        self._ov_card_table = self._make_table(["Status", "Anzahl", "Anteil"])
        right.addWidget(self._ov_card_table)
        tables_row.addLayout(right, 2)

        layout.addLayout(tables_row, 1)
        return w

    def _on_ov_alltime_changed(self):
        all_time = self._ov_all_time.isChecked()
        self._ov_from.setEnabled(not all_time)
        self._ov_to.setEnabled(not all_time)

    def _run_overview(self):
        loc_id = self._location_combo.currentData()
        if self._ov_all_time.isChecked():
            date_from = date_to = None
        else:
            date_from = self._ov_from.date().toString("yyyy-MM-dd")
            date_to   = self._ov_to.date().toString("yyyy-MM-dd")

        data = self.report_service.get_status_overview(
            location_id=loc_id,
            date_from=date_from,
            date_to=date_to,
        )

        # KPIs aktualisieren
        self._set_kpi(self._kpi_total_claims,  str(data["total_claims"]))
        self._set_kpi(self._kpi_total_cards,   str(data["cards_active"]))
        self._set_kpi(self._kpi_total_persons, str(data["total_persons"]))
        in_pruefung = next(
            (r["count"] for r in data["claim_status_rows"] if r["status_key"] == "IN_PRUEFUNG"), 0
        )
        self._set_kpi(self._kpi_open, str(in_pruefung))

        # Antrags-Tabelle
        self._ov_claim_table.setRowCount(0)
        for row in data["claim_status_rows"]:
            if row["count"] == 0:
                continue
            r = self._ov_claim_table.rowCount()
            self._ov_claim_table.insertRow(r)
            self._ov_claim_table.setItem(r, 0, _make_item(row["status"]))
            self._ov_claim_table.setItem(r, 1, _make_item(row["count"], align_right=True))
            self._ov_claim_table.setItem(r, 2, _make_item(
                _pct(row["count"], data["total_claims"]), align_right=True))
        self._append_totals_row(self._ov_claim_table, "GESAMT", data["total_claims"], 3)

        # Karten-Tabelle
        self._ov_card_table.setRowCount(0)
        for row in data["card_status_rows"]:
            if row["count"] == 0:
                continue
            r = self._ov_card_table.rowCount()
            self._ov_card_table.insertRow(r)
            self._ov_card_table.setItem(r, 0, _make_item(row["status"]))
            self._ov_card_table.setItem(r, 1, _make_item(row["count"], align_right=True))
            self._ov_card_table.setItem(r, 2, _make_item(
                _pct(row["count"], data["total_cards"]), align_right=True))
        self._append_totals_row(self._ov_card_table, "GESAMT", data["total_cards"], 3)

    # ── Tab 2: Jahresauswertung ───────────────────────────────────────────────

    def _build_annual_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)
        filter_row.addWidget(QLabel("Jahr:"))
        self._year_spin = QSpinBox()
        self._year_spin.setRange(2015, 2099)
        self._year_spin.setValue(date.today().year)
        self._year_spin.setFixedWidth(80)
        filter_row.addWidget(self._year_spin)
        self._show_prev_year = QCheckBox("Vorjahresvergleich anzeigen")
        self._show_prev_year.setChecked(True)
        filter_row.addWidget(self._show_prev_year)
        filter_row.addStretch()
        btn_annual = QPushButton("Jahresauswertung starten")
        btn_annual.setObjectName("PrimaryButton")
        btn_annual.clicked.connect(self._run_annual)
        filter_row.addWidget(btn_annual)
        layout.addLayout(filter_row)

        # Jahres-Tabelle (scrollbar)
        self._annual_table = self._make_table([
            "Monat", "Neue Anträge", "Prüfungen", "Anspruchsber.", "Härtefall",
            "Abgelehnt", "Neue Karten", "Neue Personen",
        ])
        self._annual_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self._annual_table, 1)

        # Hinweis-Label
        self._annual_hint = QLabel("")
        self._annual_hint.setStyleSheet("font-size: 11px; color: #888;")
        layout.addWidget(self._annual_hint)
        return w

    def _run_annual(self):
        year   = self._year_spin.value()
        loc_id = self._location_combo.currentData()
        data   = self.report_service.get_annual_report(year=year, location_id=loc_id)

        self._annual_table.setRowCount(0)

        prev_rows = (data.get("prev_year") or {}).get("rows") or []
        show_prev = self._show_prev_year.isChecked() and bool(prev_rows)

        _COLS_KEYS = ["new_claims", "evaluated", "approved", "hardship", "rejected", "new_cards", "new_persons"]

        for idx, row in enumerate(data["rows"]):
            r = self._annual_table.rowCount()
            self._annual_table.insertRow(r)
            self._annual_table.setItem(r, 0, _make_item(row["month_name"], bold=True))

            for col, key in enumerate(_COLS_KEYS, start=1):
                val = row[key]
                item = _make_item(val, align_right=True)
                # Grünliche Hinterlegung für genehmigte Ansprüche
                if key == "approved" and val > 0:
                    item.setBackground(_COL_APPROVED)
                elif key == "hardship" and val > 0:
                    item.setBackground(_COL_HARDSHIP)
                elif key == "rejected" and val > 0:
                    item.setBackground(_COL_REJECTED)
                self._annual_table.setItem(r, col, item)

        # Summenzeile
        t = data["totals"]
        r = self._annual_table.rowCount()
        self._annual_table.insertRow(r)
        self._annual_table.setItem(r, 0, _make_item("GESAMT", bold=True, bg=_COL_TOTALS_BG))
        for col, key in enumerate(_COLS_KEYS, start=1):
            self._annual_table.setItem(r, col, _make_item(t[key], align_right=True, bold=True, bg=_COL_TOTALS_BG))

        # Vorjahreszählen
        if show_prev:
            prev = data["prev_year"]
            r = self._annual_table.rowCount()
            self._annual_table.insertRow(r)
            prev_t = prev
            self._annual_table.setItem(r, 0, _make_item(f"Vorjahr {prev['year']}", bold=True))
            for col, key in enumerate(["new_claims", None, "approved", None, None, "new_cards", None], start=1):
                if key:
                    self._annual_table.setItem(r, col, _make_item(prev_t[key], align_right=True))

            # Delta-Zeile
            r = self._annual_table.rowCount()
            self._annual_table.insertRow(r)
            self._annual_table.setItem(r, 0, _make_item("Δ Vorjahr", bold=True))
            for col, key in enumerate(["new_claims", None, "approved", None, None, "new_cards", None], start=1):
                if key and key in prev_t:
                    delta = _delta_str(t[key], prev_t[key])
                    fg = QColor("#1a5c37") if not delta.startswith("–") and "+" in delta else None
                    if delta.startswith("-"):
                        fg = QColor("#c0392b")
                    self._annual_table.setItem(r, col, _make_item(delta, align_right=True, fg=fg))

        loc_name = data["location_name"]
        self._annual_hint.setText(
            f"Jahresauswertung {year} | {loc_name} | "
            f"{data['totals']['new_claims']} Anträge · "
            f"{data['totals']['approved']} Anspruchsberechtigt · "
            f"{data['totals']['new_cards']} Karten ausgegeben"
        )

    # ── Tab 3: Zeitraum ───────────────────────────────────────────────────────

    def _build_period_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)
        filter_row.addWidget(QLabel("Von:"))
        self._pd_from = QDateEdit()
        self._pd_from.setCalendarPopup(True)
        self._pd_from.setDate(QDate(date.today().year, 1, 1))
        self._pd_from.setDisplayFormat("dd.MM.yyyy")
        filter_row.addWidget(self._pd_from)
        filter_row.addWidget(QLabel("Bis:"))
        self._pd_to = QDateEdit()
        self._pd_to.setCalendarPopup(True)
        self._pd_to.setDate(QDate.currentDate())
        self._pd_to.setDisplayFormat("dd.MM.yyyy")
        filter_row.addWidget(self._pd_to)
        filter_row.addStretch()
        btn_pd = QPushButton("Zeitraum auswerten")
        btn_pd.setObjectName("PrimaryButton")
        btn_pd.clicked.connect(self._run_period)
        filter_row.addWidget(btn_pd)
        layout.addLayout(filter_row)

        # Kompakt-Zusammenfassung
        self._pd_summary = QLabel("")
        self._pd_summary.setStyleSheet(
            "background: #f0f8ff; border: 1px solid #b8d4f0; border-radius: 6px; "
            "padding: 8px 14px; font-size: 12px; color: #1a3a5c;"
        )
        self._pd_summary.setVisible(False)
        layout.addWidget(self._pd_summary)

        # Monatsaufschlüsselungs-Tabelle
        layout.addWidget(QLabel("<b>Monatsaufschlüsselung im Zeitraum</b>"))
        self._pd_table = self._make_table([
            "Monat", "Neue Anträge", "Prüfungen", "Anspruchsber.",
            "Härtefall", "Abgelehnt", "Neue Karten", "Neue Personen",
        ])
        self._pd_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self._pd_table, 1)

        # Vorjahreszeilenvergleich
        layout.addWidget(QLabel("<b>Vergleich mit Vorjahreszeitraum</b>"))
        self._pd_comp_table = self._make_table([
            "Kennzahl", "Aktueller Zeitraum", "Vorjahreszeitraum", "Differenz",
        ])
        self._pd_comp_table.setMaximumHeight(180)
        self._pd_comp_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self._pd_comp_table)
        return w

    def _run_period(self):
        loc_id     = self._location_combo.currentData()
        date_from  = self._pd_from.date().toString("yyyy-MM-dd")
        date_to    = self._pd_to.date().toString("yyyy-MM-dd")

        data = self.report_service.get_period_monthly_breakdown(
            start_date=date_from, end_date=date_to, location_id=loc_id
        )

        t = data["totals"]
        self._pd_summary.setText(
            f"Zeitraum: {data['period']}  |  Standort: {data['location_name']}  |  "
            f"Neue Anträge: {t['new_claims']}  |  Prüfungen: {t['evaluated']}  |  "
            f"Anspruchsberechtigt: {t['approved']}  |  Karten ausgegeben: {t['new_cards']}  |  "
            f"Neue Personen: {t['new_persons']}"
        )
        self._pd_summary.setVisible(True)

        _KEYS = ["new_claims", "evaluated", "approved", "hardship", "rejected", "new_cards", "new_persons"]

        self._pd_table.setRowCount(0)
        for row in data["rows"]:
            r = self._pd_table.rowCount()
            self._pd_table.insertRow(r)
            self._pd_table.setItem(r, 0, _make_item(row["label"], bold=True))
            for col, key in enumerate(_KEYS, start=1):
                val = row[key]
                item = _make_item(val, align_right=True)
                if key == "approved"  and val > 0: item.setBackground(_COL_APPROVED)
                if key == "hardship"  and val > 0: item.setBackground(_COL_HARDSHIP)
                if key == "rejected"  and val > 0: item.setBackground(_COL_REJECTED)
                self._pd_table.setItem(r, col, item)

        r = self._pd_table.rowCount()
        self._pd_table.insertRow(r)
        self._pd_table.setItem(r, 0, _make_item("GESAMT", bold=True, bg=_COL_TOTALS_BG))
        for col, key in enumerate(_KEYS, start=1):
            self._pd_table.setItem(r, col, _make_item(t[key], align_right=True, bold=True, bg=_COL_TOTALS_BG))

        # Vorjahresvergleich
        self._pd_comp_table.setRowCount(0)
        prev = data.get("prev_summary")
        comp_rows = [
            ("Neue Anträge",        t["new_claims"],  "total_applications"),
            ("Geprüfte Anträge",    t["evaluated"],   "total_evaluations"),
            ("Anspruchsberechtigt", t["approved"],    "approved_claims"),
            ("Abgelehnt",           t["rejected"],    "rejected_claims"),
            ("Härtefall",           t["hardship"],    "hardship_claims"),
        ]
        for label, curr_val, prev_key in comp_rows:
            prev_val = (prev or {}).get(prev_key, 0) if prev else None
            r = self._pd_comp_table.rowCount()
            self._pd_comp_table.insertRow(r)
            self._pd_comp_table.setItem(r, 0, _make_item(label))
            self._pd_comp_table.setItem(r, 1, _make_item(curr_val, align_right=True))
            prev_text  = str(prev_val) if prev_val is not None else "–"
            delta_text = _delta_str(curr_val, prev_val) if prev_val else "–"
            self._pd_comp_table.setItem(r, 2, _make_item(prev_text, align_right=True))
            self._pd_comp_table.setItem(r, 3, _make_item(delta_text, align_right=True))

    # ── Tab 4: Warteliste ─────────────────────────────────────────────────────

    def _build_waitlist_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        filter_row = QHBoxLayout()
        filter_row.addStretch()
        btn_wl = QPushButton("Warteliste aktualisieren")
        btn_wl.setObjectName("PrimaryButton")
        btn_wl.clicked.connect(self._run_waitlist)
        filter_row.addWidget(btn_wl)
        layout.addLayout(filter_row)

        info = QLabel(
            "Zeigt alle offenen Fälle (Status: In Prüfung) sortiert nach Wartezeit. "
            "Lange Wartezeiten können auf Bearbeitungsrückstände hinweisen."
        )
        info.setWordWrap(True)
        info.setStyleSheet("font-size: 12px; color: #555;")
        layout.addWidget(info)

        self._wl_table = self._make_table([
            "Fallnummer", "Person", "Standort", "Eingangsdatum", "Wartezeit (Tage)",
        ])
        self._wl_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self._wl_table, 1)
        return w

    def _run_waitlist(self):
        loc_id = self._location_combo.currentData()
        rows   = self.report_service.get_waitlist_report(location_id=loc_id)
        self._wl_table.setRowCount(0)
        for row in rows:
            r = self._wl_table.rowCount()
            self._wl_table.insertRow(r)
            self._wl_table.setItem(r, 0, _make_item(row.get("case_number", "–")))
            self._wl_table.setItem(r, 1, _make_item(row.get("person", "–")))
            self._wl_table.setItem(r, 2, _make_item(row.get("location", "–")))
            self._wl_table.setItem(r, 3, _make_item(row.get("created_at", "–")))
            days = row.get("wait_days", 0)
            days_item = _make_item(str(days), align_right=True,
                                   fg=QColor("#c0392b") if days > 30 else None)
            self._wl_table.setItem(r, 4, days_item)

    # ── Hilfsmethoden ─────────────────────────────────────────────────────────

    def _load_locations(self):
        self._locations = self.report_service.get_locations(include_inactive=False)
        self._location_combo.blockSignals(True)
        self._location_combo.clear()
        self._location_combo.addItem("Alle Standorte", None)
        for loc in self._locations:
            self._location_combo.addItem(loc["name"], loc["id"])

        if not Session.is_admin():
            loc_id = Session.get_location_id()
            for i in range(self._location_combo.count()):
                if self._location_combo.itemData(i) == loc_id:
                    self._location_combo.setCurrentIndex(i)
                    break
            self._location_combo.setEnabled(False)

        self._location_combo.blockSignals(False)

    @staticmethod
    def _make_table(headers: list[str]) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tbl.setAlternatingRowColors(True)
        tbl.verticalHeader().hide()
        return tbl

    @staticmethod
    def _append_totals_row(table: QTableWidget, label: str, total: int, ncols: int):
        r = table.rowCount()
        table.insertRow(r)
        table.setItem(r, 0, _make_item(label, bold=True, bg=_COL_TOTALS_BG))
        table.setItem(r, 1, _make_item(str(total), align_right=True, bold=True, bg=_COL_TOTALS_BG))
        for col in range(2, ncols):
            table.setItem(r, col, _make_item("", bg=_COL_TOTALS_BG))

    @staticmethod
    def _make_kpi(label: str, value: str) -> QFrame:
        card = QFrame()
        card.setObjectName("Card")
        card.setMinimumWidth(130)
        card.setMaximumWidth(200)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(4)
        lbl = QLabel(label)
        lbl.setStyleSheet("font-size: 11px; color: #888;")
        val = QLabel(value)
        val.setObjectName("_kpiValue")
        val.setStyleSheet("font-size: 22px; font-weight: bold; color: #1a3a5c;")
        layout.addWidget(lbl)
        layout.addWidget(val)
        return card

    @staticmethod
    def _set_kpi(card: QFrame, value: str):
        val_lbl = card.findChild(QLabel, "_kpiValue")
        if val_lbl:
            val_lbl.setText(value)

    # ── Export ────────────────────────────────────────────────────────────────

    def _get_active_context(self) -> tuple[QTableWidget, str, str]:
        """
        Gibt (Tabelle, Titel, Kontext-Zeile) für den aktuell sichtbaren Tab zurück.
        Kontext-Zeile enthält z. B. Standort + Jahr/Zeitraum.
        """
        idx = self._tabs.currentIndex()
        loc = self._location_combo.currentText()

        if idx == 0:
            table = self._ov_claim_table
            title = "Übersicht – Anträge nach Status"
            ctx   = f"Standort: {loc}"
        elif idx == 1:
            table = self._annual_table
            year  = self._year_spin.value()
            title = f"Jahresauswertung {year}"
            ctx   = f"Standort: {loc} | Jahr: {year}"
        elif idx == 2:
            table = self._pd_table
            d_from = self._pd_from.date().toString("dd.MM.yyyy")
            d_to   = self._pd_to.date().toString("dd.MM.yyyy")
            title  = f"Zeitraum-Auswertung {d_from} – {d_to}"
            ctx    = f"Standort: {loc} | Von: {d_from} | Bis: {d_to}"
        else:
            table = self._wl_table
            title = "Warteliste – Offene Fälle"
            ctx   = f"Standort: {loc}"

        return table, title, ctx

    def _export_to_pdf(self) -> None:
        table, title, ctx = self._get_active_context()
        if table.rowCount() == 0:
            QMessageBox.information(self, "Kein Inhalt",
                                    "Bitte zuerst eine Auswertung starten.")
            return

        default_name = f"{title.replace(' ', '_')}_{date.today()}.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF exportieren", default_name, "PDF-Dateien (*.pdf)"
        )
        if not path:
            return

        try:
            self._write_pdf(path, table, title, ctx)
            if QMessageBox.question(
                self, "Export erfolgreich",
                f"PDF gespeichert unter:\n{path}\n\nJetzt öffnen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes:
                QDesktopServices.openUrl(QUrl.fromLocalFile(path))
        except Exception as exc:
            QMessageBox.critical(self, "Exportfehler", f"PDF-Export fehlgeschlagen:\n{exc}")

    def _write_pdf(self, path: str, table: QTableWidget, title: str, ctx: str) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
        )

        ncols = table.columnCount()
        use_landscape = ncols > 6
        pagesize = landscape(A4) if use_landscape else A4
        margin = 2.0 * cm

        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(
            "ExportTitle", parent=styles["Heading1"],
            fontSize=16, leading=20, textColor=colors.HexColor("#1a3a5c"),
            fontName="Helvetica-Bold",
        )
        style_ctx = ParagraphStyle(
            "ExportCtx", parent=styles["Normal"],
            fontSize=9, textColor=colors.gray, fontName="Helvetica",
        )
        style_cell = ParagraphStyle(
            "Cell", parent=styles["Normal"],
            fontSize=8, leading=11, fontName="Helvetica",
        )

        doc = SimpleDocTemplate(path, pagesize=pagesize,
                                leftMargin=margin, rightMargin=margin,
                                topMargin=margin, bottomMargin=margin)

        story = []
        story.append(Paragraph(title, style_title))
        story.append(Paragraph(ctx, style_ctx))
        story.append(Paragraph(f"Exportiert am {date.today().strftime('%d.%m.%Y')}", style_ctx))
        story.append(HRFlowable(width="100%", thickness=1,
                                 color=colors.HexColor("#2c5f8a"), spaceAfter=10))

        # Tabellendaten aus QTableWidget extrahieren
        headers = [table.horizontalHeaderItem(c).text()
                   if table.horizontalHeaderItem(c) else f"Sp.{c+1}"
                   for c in range(ncols)]
        data = [headers]
        for r in range(table.rowCount()):
            row = []
            for c in range(ncols):
                item = table.item(r, c)
                row.append(Paragraph(item.text() if item else "", style_cell))
            data.append(row)

        avail_w = (landscape(A4)[0] if use_landscape else A4[0]) - 2 * margin
        col_w = avail_w / ncols

        tbl = Table(data, colWidths=[col_w] * ncols, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1a3a5c")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (0, -1),  "LEFT"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ]))
        story.append(tbl)
        doc.build(story)

    def _export_to_excel(self) -> None:
        table, title, ctx = self._get_active_context()
        if table.rowCount() == 0:
            QMessageBox.information(self, "Kein Inhalt",
                                    "Bitte zuerst eine Auswertung starten.")
            return

        default_name = f"{title.replace(' ', '_')}_{date.today()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel exportieren", default_name, "Excel-Dateien (*.xlsx)"
        )
        if not path:
            return

        try:
            self._write_excel(path, table, title, ctx)
            if QMessageBox.question(
                self, "Export erfolgreich",
                f"Excel-Datei gespeichert unter:\n{path}\n\nJetzt öffnen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes:
                QDesktopServices.openUrl(QUrl.fromLocalFile(path))
        except Exception as exc:
            QMessageBox.critical(self, "Exportfehler",
                                 f"Excel-Export fehlgeschlagen:\n{exc}")

    def _write_excel(self, path: str, table: QTableWidget, title: str, ctx: str) -> None:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel-Limit für Blattnamen

        # Metadaten-Zeilen
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="1A3A5C")
        ws["A2"] = ctx
        ws["A2"].font = Font(size=9, color="888888")
        ws["A3"] = f"Exportiert am {date.today().strftime('%d.%m.%Y')}"
        ws["A3"].font = Font(size=9, color="888888")
        ws.append([])

        # Styles
        header_font    = Font(bold=True, color="FFFFFF", size=10)
        header_fill    = PatternFill("solid", fgColor="1A3A5C")
        header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        totals_fill    = PatternFill("solid", fgColor="DCE9F5")
        totals_font    = Font(bold=True, size=10)
        alt_fill       = PatternFill("solid", fgColor="F5F5F5")
        thin_border    = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        ncols = table.columnCount()
        header_row_idx = ws.max_row + 1

        # Kopfzeile
        headers = [table.horizontalHeaderItem(c).text()
                   if table.horizontalHeaderItem(c) else f"Sp.{c+1}"
                   for c in range(ncols)]
        ws.append(headers)
        for cell in ws[header_row_idx]:
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Datenzeilen
        for r in range(table.rowCount()):
            row_data = []
            for c in range(ncols):
                item = table.item(r, c)
                text = item.text() if item else ""
                # Zahlen als int/float speichern für korrekte Excel-Sortierung
                try:
                    if "." in text or "," in text:
                        val = float(text.replace(",", ".").replace("%", "").strip())
                    else:
                        val = int(text)
                except (ValueError, AttributeError):
                    val = text
                row_data.append(val)
            ws.append(row_data)

            excel_row = ws.max_row
            is_totals = (r == table.rowCount() - 1 and
                         table.item(r, 0) and
                         "GESAMT" in (table.item(r, 0).text() or ""))
            for c_idx, cell in enumerate(ws[excel_row], start=1):
                cell.border = thin_border
                if is_totals:
                    cell.fill = totals_fill
                    cell.font = totals_font
                elif r % 2 == 1:
                    cell.fill = alt_fill
                if c_idx > 1:
                    cell.alignment = Alignment(horizontal="right")

        # Spaltenbreiten automatisch anpassen
        for col_cells in ws.iter_cols(min_row=header_row_idx,
                                       max_row=ws.max_row):
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)

        # Zeile fixieren (ab Kopfzeile)
        ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)
        wb.save(path)
