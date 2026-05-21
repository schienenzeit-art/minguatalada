from __future__ import annotations

from datetime import date
from pathlib import Path

from app.config import DATA_DIR
from openpyxl import Workbook
from openpyxl.styles import Font

from services.report_service import ReportService

EXCEL_ROOT = Path(DATA_DIR) / "excel"


class ExcelService:
    def __init__(self, report_service: ReportService | None = None):
        self.report_service = report_service or ReportService()
        EXCEL_ROOT.mkdir(parents=True, exist_ok=True)

    def generate_customer_report_excel(
        self,
        location_id: int | None,
        year: int,
        month: int,
        file_path: str | None = None,
    ) -> str:
        report = self.report_service.get_customer_monthly_report(location_id, year, month)

        if file_path is None:
            filename = f"Kundenbestand_{report['location_name'].replace(' ', '_')}_{year}_{month:02d}.xlsx"
            file_path = str(EXCEL_ROOT / filename)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Monatsreport"

        header_font = Font(bold=True)
        sheet["A1"] = "Bericht"
        sheet["B1"] = "Wert"
        sheet["A1"].font = header_font
        sheet["B1"].font = header_font

        rows = [
            ["Standort", report["location_name"]],
            ["Berichtsmonat", report["report_month"]],
            ["Neukunden im Monat", report["new_customers"]],
            ["Gesamtbestand Kunden", report["total_customers"]],
        ]

        for index, row in enumerate(rows, start=2):
            sheet[f"A{index}"] = row[0]
            sheet[f"B{index}"] = row[1]

        workbook.save(file_path)
        return str(Path(file_path).resolve())

    def generate_period_report_excel(
        self,
        location_id: int | None,
        start_date: str,
        end_date: str,
        selected_metrics: list[str] | None = None,
        file_path: str | None = None,
    ) -> str:
        report = self.report_service.get_period_report(start_date, end_date, location_id)
        if file_path is None:
            location_part = report["location_name"].replace(" ", "_")
            file_path = str(EXCEL_ROOT / f"Zeitraumreport_{location_part}_{start_date}_{end_date}.xlsx")

        if selected_metrics is None:
            selected_metrics = [
                "applications",
                "evaluations",
                "approved",
                "rejected",
                "hardship",
                "cards",
            ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bericht"

        header_font = Font(bold=True)
        sheet["A1"] = "Bericht"
        sheet["B1"] = "Wert"
        sheet["A1"].font = header_font
        sheet["B1"].font = header_font

        rows = [
            ["Standort", report["location_name"]],
            ["Zeitraum", f"{start_date} bis {end_date}"],
        ]

        if "applications" in selected_metrics:
            rows.append(["Anträge", report["total_applications"]])
        if "evaluations" in selected_metrics:
            rows.append(["Prüfungen", report["total_evaluations"]])
        if "approved" in selected_metrics:
            rows.append(["anspruchsberechtigte", report["approved_claims"]])
        if "rejected" in selected_metrics:
            rows.append(["abgelehnte", report["rejected_claims"]])
        if "hardship" in selected_metrics:
            rows.append(["Härtefälle", report["hardship_claims"]])

        for index, row in enumerate(rows, start=2):
            sheet[f"A{index}"] = row[0]
            sheet[f"B{index}"] = row[1]

        if "cards" in selected_metrics and report["cards_by_location"]:
            cards_sheet = workbook.create_sheet(title="Kundenstamm")
            cards_sheet["A1"] = "Standort"
            cards_sheet["B1"] = "Karten im Zeitraum"
            cards_sheet["A1"].font = header_font
            cards_sheet["B1"].font = header_font
            for row_index, item in enumerate(report["cards_by_location"], start=2):
                cards_sheet[f"A{row_index}"] = item.get("location_name", "-")
                cards_sheet[f"B{row_index}"] = item.get("card_count", 0)

        workbook.save(file_path)
        return str(Path(file_path).resolve())
